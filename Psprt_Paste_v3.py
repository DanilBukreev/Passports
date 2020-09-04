#!/usr/bin/python3
# -*- coding: utf-8 -*-

from __future__ import print_function
from mailmerge import MailMerge
from openpyxl import Workbook,load_workbook
import PySimpleGUI as sg
from docx import Document
import os
from win32com.client import Dispatch as comDispatch


def multiple_replace(target_str, replace_values):
    # получаем заменяемое: подставляемое из словаря в цикле
    for i, j in replace_values.items():
        # меняем все target_str на подставляемое
        target_str = target_str.replace(i, j)
    return target_str


def all():
    sg.theme('DarkBlue12')
    #sg.theme('Green')
    #sg.theme_previewer()
    layout2 = [
          [sg.Text('Укажите расположение до Word-шаблона:', size=(35, 1)), sg.InputText(size=(80, 1)), sg.FileBrowse()],
          [sg.Text('Укажите расположение файла Excel:', size=(35, 1)), sg.InputText(size=(80, 1)), sg.FileBrowse()],
          [sg.Text('Укажите путь для сохранения Паспортов:', size=(35, 1)), sg.InputText(size=(80, 1)), sg.FolderBrowse()],
          [sg.Submit(), sg.Cancel()]]
    window2 = sg.Window('Меню', layout2)
    event, values = window2.read()
    window2.close()
    WORD_path, Excel_path, Save_path = values[0], values[1], values[2]


    wb = load_workbook(Excel_path)
    sheet = wb.worksheets[0]
    for cell in sheet["A"]:
        if cell.value is None:
            cl= cell.row
            #print(cl)
            break

    short_asu = []
    for val in range (4,cl):
        short_asu.append(sheet.cell(row=val, column=3).value)
    replace_values = {' ': "_", '"': "_", '/': '_'}


    sg.theme('DarkBlue12')
    progressbar = [
        [sg.ProgressBar(len(short_asu), orientation='h', size=(78, 10), key='progressbar')]
    ]
    outputwin = [
        [sg.Output(size=(119,20))]
    ]

    layout = [
        [sg.Frame('Progress',layout= progressbar)],
        [sg.Frame('Output', layout = outputwin)],
        [sg.Submit('Start'),sg.Cancel()]
    ]

    window = sg.Window('Custom Progress Meter', layout)
    progress_bar = window['progressbar']
    while True:
        event, values = window.read()
        if event == 'Cancel'  or event is None:
            raise SystemExit(1)


        elif event == 'Start':
            print("****** START ******")
            try:
                files = os.listdir(Save_path)
                if not files:
                    os.mkdir(Save_path + '/' + 'Шаблоны_паспортов')
                info_dir = 0
                for f, n in enumerate(files):
                    if n == 'Шаблоны_паспортов':
                        info_dir = 1
                    else:
                        info_dir = 2
                if info_dir == 2:
                    os.mkdir(Save_path + '/' + 'Шаблоны_паспортов')
                for d, item in enumerate(short_asu, 4):
                    myASU = []
                    if item == 'Исключено':
                        continue
                    for val in range(1, 50):
                        myASU.append(sheet.cell(row=d, column=val).value)
                    replace_str = multiple_replace(myASU[2], replace_values)
                    document = MailMerge(WORD_path)
                    # print(document.get_merge_fields())
                    # print(myASU)
                    document.merge(
                        Полное_наим=str(myASU[1]),
                        Краткое_наим=str(myASU[2]),
                        Собственник_АСУ_ТП=str(myASU[3]),
                        Эксп_Орг=str(myASU[4]),
                        Назначение_п1_3=str(myASU[5]),
                        адрес_асу=str(myASU[6]),
                        Владелец_АСУТП=str(myASU[7]),
                        п1_6=str(myASU[8]),
                        Класс_Опасности=str(myASU[9]),
                        Крит_Тех_Проц=str(myASU[10]),
                        Соц_знач=str(myASU[11]),
                        Эконом_знач=str(myASU[12]),
                        Эколог_знач=str(myASU[13]),
                        п1_10=str(myASU[14]),
                        Режим_работы_АСУ_ТП=str(myASU[15]),
                        Наим_Тех_проц=str(myASU[16]),
                        Опис_арх_асу=str(myASU[17]),
                        Описание_п3_1=str(myASU[18]),
                        Описание_п3_2=str(myASU[19]),
                        Описание_п3_3=str(myASU[20]),
                        п3_7=str(myASU[21]),
                        Идент_Аутент=str(myASU[22]),
                        Описание_табл_п5_1=str(myASU[23]),
                        Упр_Доступом=str(myASU[24]),
                        Огрн_прог_среды=str(myASU[25]),
                        Защита_маш_нос_инф=str(myASU[26]),
                        Ауд_ИБ=str(myASU[27]),
                        Антивир=str(myASU[28]),
                        Пред_Вторж=str(myASU[29]),
                        Целостность=str(myASU[30]),
                        Резерв_оборуд=str(myASU[31]),
                        Рез_Коп=str(myASU[32]),
                        ЗИП=str(myASU[33]),
                        Мон_Тех_Сост=str(myASU[34]),
                        п5_10=str(myASU[35]),
                        Меры_физ_защ1=str(myASU[36]),
                        Меры_физ_защ2=str(myASU[37]),
                        Меры_физ_защ3=str(myASU[38]),
                        Меры_физ_защ4=str(myASU[39]),
                        Меры_физ_защ5=str(myASU[40]),
                        ИБП=str(myASU[41]),
                        п5_11=str(myASU[42]),
                        п5_12=str(myASU[43]),
                        У_Конфиг=str(myASU[44]),
                        п5_14=str(myASU[45]),
                        Реаг_Инц_ИБ=str(myASU[46]),
                        п6_16=str(myASU[47]),
                        Инф_обуч_персн=str(myASU[48])
                    )
                    document.write(Save_path + '/' + str(myASU[0]) + '_' + 'Паспорт_' + replace_str + '.docx')
                    dubl = Document(WORD_path)
                    fulPath = Save_path + '/' + 'Шаблоны_паспортов/' + 'Шаблон_' + str(myASU[0]) + '_' + replace_str + '.docx'
                    dubl.save(fulPath)
                    word = comDispatch('Word.Application')
                    doc = word.Documents.Open(Save_path + '/' + str(myASU[0]) + '_' + 'Паспорт_' + replace_str + '.docx')
                    word.Application.Run(Save_path + '/' + str(myASU[0]) + '_' + 'Паспорт_' + replace_str + '.docx' + '!MakeList')  # возможно здесь тоже нужен полный путь
                    doc.Save()
                    doc.Close()
                    print(str(myASU[0]) + ") ", myASU[2], " - Done")
                    progress_bar.UpdateBar(d + 1)
                print("****** FINISH ******")
            except FileNotFoundError:
                print("*****Невозможно сохранить файл. ОШИБКА В НАЗВАНИИ КРАТКОГО АСУ. ИСПОЛЬЗОВАНЫ НЕДОПУСТИМЫЕ СИМВОЛЫ "" или /   ******")
            except Exception as e:
                print(e.message, e.args)




def some():
    sg.theme('DarkBlue12')
    # sg.theme('Green')
    # sg.theme_previewer()
    layout2 = [
        [sg.Text('Укажите путь до папки Шаблоны_Паспортов:', size=(35, 1)), sg.InputText(size=(80, 1)), sg.FolderBrowse()],
        [sg.Text('Укажите расположение файла Excel:', size=(35, 1)), sg.InputText(size=(80, 1)), sg.FileBrowse()],
        [sg.Text('Укажите путь для сохранения Паспортов:', size=(35, 1)), sg.InputText(size=(80, 1)),
         sg.FolderBrowse()],
        [sg.Submit(), sg.Cancel()]]
    window2 = sg.Window('Меню', layout2)
    event, values = window2.read()
    window2.close()
    WORD_path, Excel_path, Save_path = values[0], values[1], values[2]

    files = os.listdir(Save_path)
    sg.theme('DarkBlue12')
    progressbar = [
        [sg.ProgressBar(len(files), orientation='h', size=(78, 10), key='progressbar')]
    ]
    outputwin = [
        [sg.Output(size=(119, 20))]
    ]

    layout = [
        [sg.Frame('Progress', layout=progressbar)],
        [sg.Frame('Output', layout=outputwin)],
        [sg.Submit('Start'), sg.Cancel()]
    ]

    window = sg.Window('Custom Progress Meter', layout)
    progress_bar = window['progressbar']
    replace_values = {' ': "_", '"': "_", '/': '_'}
    while True:
        event, values = window.read(timeout=10)
        if event == 'Cancel' or event is None:
            exit()
            break
        elif event == 'Start':
            print("****** START ******")
            try:
                files_templates = os.listdir(Save_path + '/' + 'Шаблоны_паспортов')
                if not files_templates:
                    print("Отсутствуют файлы шаблонов в указанной папке")
                if not files:
                    os.mkdir(Save_path + '/' + 'Шаблоны_паспортов')
                info_dir = 0
                for f, n in enumerate(files):
                    if n == 'Шаблоны_паспортов':
                        info_dir = 1
                    else:
                        info_dir = 2
                if info_dir == 2:
                    os.mkdir(Save_path + '/' + 'Шаблоны_паспортов')

                wb = load_workbook(Excel_path)
                sheet = wb.worksheets[0]
                for cell in sheet["A"]:
                    if cell.value is None:
                        cl = cell.row
                        # print(cl)
                        break

                directory = WORD_path
                files = os.listdir(directory)  # Массив из всех документов в папке, которые будут использоваться как шаблон
                for d in range(0, len(files)):
                    splt_files = list(files[d])
                    # print(splt_files)
                    ful_number = ''.join([splt_files[7], splt_files[8], splt_files[9]])
                    wbSearch = load_workbook(Excel_path)
                    wsSearch = wbSearch.worksheets[0]
                    for cell in wsSearch["A"]:
                        if cell.value == ful_number:
                            cl = cell.row
                            # print(cl)
                            break
                    myASU = []
                    if files[d] == 'Исключено':
                        continue
                    for val in range(1, 50):
                        myASU.append(sheet.cell(row=cl, column=val).value)
                    replace_str = multiple_replace(myASU[2], replace_values)
                    document = MailMerge(WORD_path + '/' + files[d])
                    # print(document.get_merge_fields())
                    # print(myASU)
                    document.merge(
                        Полное_наим=str(myASU[1]),
                        Краткое_наим=str(myASU[2]),
                        Собственник_АСУ_ТП=str(myASU[3]),
                        Эксп_Орг=str(myASU[4]),
                        Назначение_п1_3=str(myASU[5]),
                        адрес_асу=str(myASU[6]),
                        Владелец_АСУТП=str(myASU[7]),
                        п1_6=str(myASU[8]),
                        Класс_Опасности=str(myASU[9]),
                        Крит_Тех_Проц=str(myASU[10]),
                        Соц_знач=str(myASU[11]),
                        Эконом_знач=str(myASU[12]),
                        Эколог_знач=str(myASU[13]),
                        п1_10=str(myASU[14]),
                        Режим_работы_АСУ_ТП=str(myASU[15]),
                        Наим_Тех_проц=str(myASU[16]),
                        Опис_арх_асу=str(myASU[17]),
                        Описание_п3_1=str(myASU[18]),
                        Описание_п3_2=str(myASU[19]),
                        Описание_п3_3=str(myASU[20]),
                        п3_7=str(myASU[21]),
                        Идент_Аутент=str(myASU[22]),
                        Описание_табл_п5_1=str(myASU[23]),
                        Упр_Доступом=str(myASU[24]),
                        Огрн_прог_среды=str(myASU[25]),
                        Защита_маш_нос_инф=str(myASU[26]),
                        Ауд_ИБ=str(myASU[27]),
                        Антивир=str(myASU[28]),
                        Пред_Вторж=str(myASU[29]),
                        Целостность=str(myASU[30]),
                        Резерв_оборуд=str(myASU[31]),
                        Рез_Коп=str(myASU[32]),
                        ЗИП=str(myASU[33]),
                        Мон_Тех_Сост=str(myASU[34]),
                        п5_10=str(myASU[35]),
                        Меры_физ_защ1=str(myASU[36]),
                        Меры_физ_защ2=str(myASU[37]),
                        Меры_физ_защ3=str(myASU[38]),
                        Меры_физ_защ4=str(myASU[39]),
                        Меры_физ_защ5=str(myASU[40]),
                        ИБП=str(myASU[41]),
                        п5_11=str(myASU[42]),
                        п5_12=str(myASU[43]),
                        У_Конфиг=str(myASU[44]),
                        п5_14=str(myASU[45]),
                        Реаг_Инц_ИБ=str(myASU[46]),
                        п6_16=str(myASU[47]),
                        Инф_обуч_персн=str(myASU[48])
                    )
                    document.write(Save_path + '/' + str(myASU[0]) + '_' + 'Паспорт_' + replace_str + '.docx')
                    word = comDispatch('Word.Application')
                    doc = word.Documents.Open(Save_path + '/' + str(myASU[0]) + '_' + 'Паспорт_' + replace_str + '.docx')
                    word.Application.Run(Save_path + '/' + str(myASU[0]) + '_' + 'Паспорт_' + replace_str + '.docx' + '!MakeList')  # возможно здесь тоже нужен полный путь
                    doc.Save()
                    doc.Close()
                    print(str(myASU[0]) + ") ", myASU[2], " - Done")
                    progress_bar.UpdateBar(d + 2)
                print("****** FINISH ******")
            except FileNotFoundError:
                print("*****Невозможно сохранить файл. ОШИБКА В НАЗВАНИИ КРАТКОГО АСУ. ИСПОЛЬЗОВАНЫ НЕДОПУСТИМЫЕ СИМВОЛЫ "" или /   ******")
            except Exception as e:
                print(e.message, e.args)
            #except:
               # print("***** SHIT,ERROR ******")

    window.close()









